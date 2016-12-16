import os
import sys
import re

from openpyxl import *
from openpyxl.styles import *

##chip = {
##    0 : PRESSO54608
##    1 : K65F180M
##    }

osList = {
    'bm' : 'KSDK_bm',
    'ksdk_bm' : 'KSDK_bm',
    'lite_bm' : 'KSDK_bm',
    'bm_release' : 'KSDK_bm',
    'freertos_release' : 'KSDK_bm',
    'freertos' : 'KSDK_freertos',
    'ksdk_freertos' : 'KSDK_freertos'
    }
testEnvList = {
    'default' : 'default_env',
    'fs'  : 'default_env',
    'hs'  : 'EHCI'
    }
moduleList = {
    'host' : 'new_usb0_host_demo',
    'dev' : 'new_usb0_device_demo',
    'device' : 'new_usb0_device_demo',
    'otg' : 'new_usb0_otg_demo',
    'other' : 'new_usb0_other_test'
    }
targetList = {
    'bm_release' : 'KSDK_bm',
    'freertos_release' : 'KSDK_freertos'
    }

def dest_excel_init(excelSour):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "cycle cases"
    ws1.append([])
    ws1.append(['','Id','Prj','Chip','Board Type','Os','IDE','Target','Test Env',\
                'Module','Testcase','B-Res','Result','Comment','CRID','Testor'])
    ws2 = wb.create_sheet(title = "binary test")
    ws2.append(['Examples','Platform','Tester'])
    destName = excelSour[:-5]+'_task.xlsx'
    wb.save(filename = destName)
    return wb

def gen_excel(excelSour):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "cycle cases"
    ws1.append([])
    ws1.append(['','Id','Prj','Chip','Board Type','Os','IDE','Target',\
                'Test Env','Module','Testcase','B-Res','Result','Comment',\
                'CRID','Testor'])
    for col in range(1,26):
        cell=ws1.cell(row=2,column=col)
        cell.font=Font(size=13,bold=True)

    ws2 = wb.create_sheet(title = "binary test")
    ws2.append(['Examples','Platform','Tester'])
    for col in range(1,26):
        cell=ws2.cell(row=1,column=col)
        cell.font=Font(size=13,bold=True)
    
    getSour = load_workbook(excelSour,read_only = True)
    getDetail = getSour.get_sheet_by_name('Details')
    count = 0
    
    col = 4
    target = getDetail.cell(row=3,column=col).value
    while (target != None):
        temp = getDetail.cell(row=1,column=col).value
        if temp != None:
            ide = temp
        #IDEList.update({col:ide})
        temp = getDetail.cell(row=2,column=col).value
        caseSuffix = ''
        if temp != None:
            if 'lite' in temp.lower():
                caseSuffix = 'lite'
            os = temp.lower()
        #IDEList.update({col:os})
        row = 4
        case = getDetail.cell(row=row,column=2).value
        while (case != None):
            if 'host' in case:
                mod = 'host'
            elif 'dev' in case:
                mod = 'dev'
            else:
                mod = 'otg'
            platform = getDetail.cell(row=row, column=col).value
            if platform != None:
                boardType = platform.split('-')[0]
                if '_' in platform:
                    chip = platform.split('-')[-1].split('_')[0]
                    mode = platform.split('-')[-1].split('_')[-1].lower()
                else:
                    chip = platform.split('-')[-1]
                    mode = 'default'
                count += 1
                ws1.append(['',str(count),chip+'-'+boardType+'-'+osList[os],chip,boardType,osList[os],ide,target,\
                    testEnvList[mode],moduleList[mod],case+caseSuffix,'','','','',''])
            row += 1
            case = getDetail.cell(row=row,column=2).value
        col += 1
        target = getDetail.cell(row=3,column=col).value
        
    while(True):
        row +=1
        temp = getDetail.cell(row=row,column=1).value
        if temp != None:
            if 'item' in temp.lower():
                break
    itemRow  = row
    row +=1
    case = getDetail.cell(row=row,column=2).value
    while (case != None):
        temp = getDetail.cell(row=row,column=1).value
        if temp != None:
            if bool(re.match('compatibility',temp,re.IGNORECASE)):
                item = 'compatibility'
            elif bool(re.match('cv_test',temp,re.IGNORECASE)):
                item = 'cv_test'
            elif bool(re.match('binary',temp,re.IGNORECASE)):
                item = 'binary'
            else:
                print 'the item cannot be recognized'
                exit(0)
            if 'host' in case:
                mod = 'host'
            elif 'dev' in case:
                mod = 'dev'
            else:
                mod = 'other'
        col = 4
        osTarget = getDetail.cell(row=itemRow,column=col).value
        while (osTarget != None):
            platform = getDetail.cell(row=row,column=col).value
            if platform != None:
                boardType = platform.split('-')[0]
                if '_' in platform:
                    chip = platform.split('-')[-1].split('_')[0]
                    mode = platform.split('-')[-1].split('_')[-1].lower()
                else:
                    chip = platform.split('-')[-1]
                    mode = 'default'
                count += 1
                if 'binary' == item:
                    ws2.append([case,chip+'-'+boardType,''])
                else:
                    ws1.append(['',str(count),chip+'-'+boardType+'-'+osList[osTarget.lower()],chip,boardType,osList[osTarget.lower()],'IAR',targetList[osTarget.lower()],\
                        testEnvList[mode],moduleList[mod],case+caseSuffix,'','','','',''])
            col += 1
            osTarget = getDetail.cell(row=itemRow,column=col).value
        row +=1
        case = getDetail.cell(row=row,column=2).value
    destName = excelSour[:-5]+'_task.xlsx'
    ws1.page_setup.fitToHeight = 0
    ws1.page_setup.fitToWidth = 1
    ws2.page_setup.fitToHeight = 0
    ws2.page_setup.fitToWidth = 1
    wb.save(filename = destName)
    return destName

if __name__ == '__main__':
    sys.stdout.flush()
    sourceFile = sys.argv[1]
    if '.xlsx' in sourceFile:
        gen_excel(sourceFile)
    else:
        print 'please input right argument(.xlsx)'
